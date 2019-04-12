# -*- coding: utf-8 -*-

__help_text__ = '''
Usage:

    python sqlite2xlsx.py input_db [recall_db] output_xlsx

Parameters:
- input_db Dataloader.to_sqlite() 导出的数据库文件
- recall_db [可选] RecallData 维护的召回数据库文件
- output_xlsx 输出的 Excel 文件

输出文件中将含有以下几张表（sheet）：
- 知识库 线上模型数据集

    | 主问题 | 标准答案 | 问法 |

- labeled RecallData 数据库中 labeled 表，包含打标后的数据

    | 问法 | 标注答案 |

- recall RecallData 数据库中 recall 表

    | 问法 |
'''

import sqlite3
import openpyxl as xl
from openpyxl.styles import Font as xlfont

import sys
import os


''' Command Line Argument Parsing '''

input_db_path, recall_db_path, output_xlsx_path = [None,] * 3

if '-h' in sys.argv or '--help' in sys.argv:
    print(__help_text__)
    exit(0)
elif len(sys.argv) == 3:
    input_db_path, output_xlsx_path = sys.argv[1:]
elif len(sys.argv) == 4:
    input_db_path, recall_db_path, output_xlsx_path = sys.argv[1:]
else:
    print('Usage: python sqlite2xlsx.py input_db [recall_db] output_xlsx')
    exit(-1)

if not os.path.exists(input_db_path):
    print('"{}" not found!'.format(input_db_path))
    exit(-1)
if os.path.exists(output_xlsx_path):
    op = input('File "{}" already exists! Overwrite? [Y/n] '.format(output_xlsx_path))
    if op.lower() in 'y' or op.lower() == 'yes':
        pass
    else:
        exit(-1)
if recall_db_path is not None and not os.path.exists(recall_db_path):
    print('Recall database at "{}" not found!'.format(recall_db_path))
    exit(-1)


''' Exporting Active Knowledge Base '''

workbook = xl.Workbook()
worksheet = workbook.active
worksheet.title = '知识库'
worksheet.cell(1, 1, value='主问题')
worksheet.cell(1, 2, value='标准答案')
worksheet.cell(1, 3, value='问法')
for cell in worksheet[1]:
    cell.font = xl.styles.Font(bold=True)

def get_first_empty_cell(ws, row: int):
    '''
    Note: 1-indexed.
    '''
    for idx, cell in enumerate(ws[row]):
        if cell.value is None:
            return idx + 1
    return len(ws[row]) + 1

with sqlite3.connect(input_db_path) as conn:
    # iter over main_ques
    cur = conn.execute('SELECT `id`, `ques`, `ans` FROM `main_ques`')
    main_ques = cur.fetchall()
    # row id in excel = id_ + 2
    for id_, ques, ans in main_ques:
        worksheet.cell(id_ + 2, 1, value=ques)
        worksheet.cell(id_ + 2, 2, value=ans)
    del main_ques
    # iter over sub_ques
    cur = conn.execute('SELECT `ques`, `main_ques_id` FROM `sub_ques`')
    sub_ques = cur.fetchall()
    for ques, main_ques_id in sub_ques:
        col_idx = get_first_empty_cell(worksheet, main_ques_id + 2)
        worksheet.cell(
            main_ques_id + 2,
            col_idx,
            value=ques
        )

workbook.save(output_xlsx_path)


''' Exporting Recall Database '''

if not recall_db_path:
    exit(0)

# export `labeled` table

worksheet = workbook.create_sheet('labeled')
worksheet.cell(1, 1, value='问法')
worksheet.cell(1, 2, value='标注答案')

with sqlite3.connect(recall_db_path) as conn:
    cur = conn.execute('SELECT `ques`, `ans` FROM `labeled`')
    labeled = cur.fetchall()
    for row_id, (ques, ans) in enumerate(labeled):
        worksheet.cell(row_id + 2, 1, value=ques)
        worksheet.cell(row_id + 2, 2, value=ans)
    del labeled

# export `recall` table

worksheet = workbook.create_sheet('recall')
worksheet.cell(1, 1, value='问法')

with sqlite3.connect(recall_db_path) as conn:
    cur = conn.execute('SELECT `ques` FROM `recall`')
    recall = cur.fetchall()
    for row_id, (ques,) in enumerate(recall):
        worksheet.cell(row_id + 2, 1, value=ques)
    del recall

workbook.save(output_xlsx_path)
